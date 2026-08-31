"""Import the user-maintained levequest route workbook into a static browser dataset."""
import json
import re
import sys
from datetime import datetime, timezone

from openpyxl import load_workbook

source, target, verification_path = sys.argv[1:4]
book = load_workbook(source, data_only=True, read_only=True)
sheet = book[book.sheetnames[0]]
jobs = {'刻木匠', '锻铁匠', '铸甲匠', '雕金匠', '制革匠', '裁衣匠', '炼金术士', '烹饪师'}
# Wiki 理符列表截图逐项复核（等级 N 的值表示 N → N+1 所需经验）。
level_experience = [0, 300, 450, 630, 970, 1440, 1940, 3000, 3920, 4970, 5900, 7430, 8620, 10200, 11300, 13100, 15200, 17400, 19600, 21900, 24300, 27400, 30600, 33900, 37300, 40800, 49200, 54600, 61900, 65600, 68400, 74000, 82700, 88700, 95000, 102000, 113000, 121000, 133000, 142000, 155000, 163000, 171000, 179000, 187000, 195000, 214000, 229000, 244000, 259000, 421000, 500000, 580000, 663000, 749000, 837000, 927000, 1019000, 1114000, 1211000, 1387000, 1456000, 1534000, 1621000, 1720000, 1834000, 1968000, 2126000, 2317000, 2550000, 2923000, 3018000, 3153000, 3324000, 3532000, 3770600, 4066000, 4377000, 4777000, 5256000, 5992000, 6171000, 6942000, 7205000, 7948000, 8287000, 9231000, 9529000, 10459000, 10838000, 13278000, 13659000, 15348000, 15912000, 17534000, 18263000, 20322000, 20957000, 22979000, 23789000]
level_experience_corrections = [
    {'level': 16, 'previous': 14200, 'corrected': 15200},
    {'level': 23, 'previous': 33000, 'corrected': 33900},
    {'level': 29, 'previous': 65000, 'corrected': 65600},
    {'level': 77, 'previous': 4387700, 'corrected': 4377000}
]
rows, job = [], None
numeric = re.compile(r'^\d+(?:-\d+)?$')

def normalize(value):
    return re.sub(r'[（）()【】\[\]\s·・，,。.!！?？：:]', '', str(value or '')).replace('制作委托', '').replace('筹集委托', '').replace('批发委托', '')

def route_key(job, level, quest, item):
    return '|'.join([job, str(level), normalize(quest), normalize(item)])

with open(verification_path, encoding='utf-8') as verification_file:
    verification_text = verification_file.read()
verification = json.loads(verification_text.split('=', 1)[1].rstrip(';\n'))
verified_routes = {entry['key']: entry for entry in verification.get('entries', [])}

for values in sheet.iter_rows(values_only=True):
    values = [str(value).strip() if value is not None else '' for value in values[:9]]
    values += [''] * (9 - len(values))
    section, _, level, item, quantity, allowances, quest, place, note = values
    if section in jobs:
        job = section
    if not job or not (numeric.fullmatch(level) and numeric.fullmatch(quantity) and numeric.fullmatch(allowances)) or not item or not quest:
        continue
    route_quantity, route_allowances = int(quantity.split('-')[0]), int(allowances.split('-')[0])
    key = route_key(job, int(level.split('-')[0]), quest, item)
    audit = verified_routes.get(key, {})
    rows.append({
        'job': job, 'level': int(level.split('-')[0]), 'item': item, 'itemId': audit.get('itemId'), 'itemIcon': audit.get('itemIcon'),
        'routeQuantity': route_quantity, 'routeAllowances': route_allowances,
        'submissionsPerAllowance': route_quantity // route_allowances if route_allowances else 0,
        'quest': quest, 'place': place, 'note': note,
        'experiencePerSubmission': audit.get('experiencePerSubmission'),
        'verified': audit.get('status') in {'garland-verified', 'wiki-manual-verified'},
        'verificationStatus': audit.get('status', 'unverified'),
        'wikiStatus': audit.get('wikiStatus', 'pending-manual-check'),
        'leveId': audit.get('leveId'),
        'wikiUrl': audit.get('wikiUrl', ''),
        'garlandUrl': audit.get('garlandUrl', ''),
        'verificationNote': audit.get('note', '')
    })

payload = {
    'schema': 1,
    'version': '0.0.1',
    'publishedAt': datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace('+00:00', 'Z'),
    'sources': {
        'routeWorkbook': '联合商城.xlsx',
        'garland': 'https://www.garlandtools.org/db/',
        'wiki': 'https://ff14.huijiwiki.com/',
        'levelExperience': '灰机 Wiki 理符列表升级经验截图（2026-08-29）'
    },
    'audit': {
        'generatedAt': verification.get('generatedAt'),
        'total': verification.get('total', len(rows)),
        'counts': verification.get('counts', {}),
        'wikiFetch': verification.get('sources', {}).get('wikiFetch', ''),
        'levelExperienceCorrections': level_experience_corrections,
        'pending': [entry for entry in verification.get('entries', []) if entry.get('status') not in {'garland-verified', 'wiki-manual-verified'}]
    },
    'jobs': ['刻木匠', '锻铁匠', '铸甲匠', '雕金匠', '制革匠', '裁衣匠', '炼金术士', '烹饪师'],
    'levelExperience': level_experience,
    'routes': rows
}
with open(target, 'w', encoding='utf-8') as output:
    output.write('// 自动生成：tools/import-levequests.py。来源：联合商城.xlsx。\n')
    output.write('window.FF14_LEVEQUESTS = ')
    json.dump(payload, output, ensure_ascii=False, separators=(',', ':'))
    output.write(';\n')
